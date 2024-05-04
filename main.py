import tkinter as tk
from tkinter import messagebox
import random
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# File paths
workers_file = 'workers_data.json'
assigned_file = 'assigned_data.xlsx'

# Check if workers data file exists
try:
    with open(workers_file, 'r') as file:
        workers = json.load(file)
except FileNotFoundError:
    # Default data if file doesn't exist
    workers = {
        'Day Shift': {
            'Alice': ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job', 'Make 301', 'Make 302',
                      'Make 303'],
            'Bob': ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job', 'Make 301', 'Make 302',
                    'Make 303'],
            'Emily': ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job', 'Make 301', 'Make 302',
                      'Make 303']
        },
        'Evening Shift': {
            'Charlie': ['Make 301', 'Make 302', 'Make 303'],
            'David': ['Make 301', 'Make 302', 'Make 303']
        }
    }

# Sample data
# Sample data
stations = ['Pak 5 (Truck)',
            'Pak 5 (Gulv)',
            'Bånde Gulv',
            'Pak totes bord 1',
            'Pak totes bord 2',
            'Pak totes bord 3',
            'Pak totes bord 4',
            'Mezzanin',
            'Pluk 1',
            'Pluk 2',
            'Pluk 3',
            'Truck job 1',
            'Truck job 2',
            'Recv 1',
            'Recv 2',
            'Decant 1',
            'Decant 2',
            'MC Recv/Decant',
            'Flow 8', 'Flow 1',
            'Flow 2 402',
            'Flow 403',
            'Flow 404',
            'Flow 3 301',
            'Flow 3 302',
            'Flow 3 303',
            'Flow 4 Maskine',
            'Flow 4 Pak 1',
            'Flow 4 Pak 2',
            'Flow 5',
            'Flow 6',
            'Flow 10',
            'Automatikmontør medhjælper',
            'Matriale / Kanban optælling',
            'Support 1',
            'Support 2',
            'Kontor',
            'Dispatch',
            'Ekstra opgave gå til TL',
            'Retur']
intervals_day_shift = ['6:00-8:00', '8:00-10:00', '10:00-12:00', '12:00-14:00']  # Updated intervals for day shift
intervals_evening_shift = ['14:00-16:00', '16-18:00', '18:00-20:00', '20:00-22:30']  # Updated intervals for evening shift


# Create the main window
window = tk.Tk()
window.title("ALFA LAVAL DC - Create Rotationsplan")

# Add padding around the elements
pad_x = 15
pad_y = 5



# Labels for the checkboxes
day_frame_label = tk.Label(window, text="Day Shift Workers")
day_frame_label.grid(row=3, column=0, padx=pad_x, pady=pad_y)

evening_frame_label = tk.Label(window, text="Evening Shift Workers")
evening_frame_label.grid(row=3, column=1, padx=0, pady=0)

# Display the checkboxes with padding
day_worker_frame = tk.Frame(window)
day_worker_frame.grid(row=5, column=0, padx=pad_x, pady=pad_y)

evening_worker_frame = tk.Frame(window)
evening_worker_frame.grid(row=5, column=1, padx=pad_x, pady=pad_y)

day_checkboxes = []
evening_checkboxes = []



# Function to handle checkbox selection
def on_checkbox_select():
    # Clear any existing buttons
    for widget in window.winfo_children():
        if isinstance(widget, tk.Button) and widget.winfo_parent() in (day_worker_frame, evening_worker_frame):
            widget.destroy()

edit_buttons_dict = {}

# Function to toggle the state of checkboxes for a given shift
def toggle_check_all_workers(shift_checkboxes):
    all_checked = all(checkbox.get() for checkbox in shift_checkboxes)
    new_state = not all_checked  # Toggle the state

    for checkbox in shift_checkboxes:
        checkbox.set(new_state)

# Button to toggle check/uncheck for all day shift workers
toggle_check_all_day_button = tk.Button(window, text="Check/Uncheck All Day Workers",
                                        command=lambda: toggle_check_all_workers(day_checkboxes))
toggle_check_all_day_button.grid(row=4, column=0, padx=pad_x, pady=pad_y)

# Button to toggle check/uncheck for all evening shift workers
toggle_check_all_evening_button = tk.Button(window, text="Check/Uncheck All Evening Workers",
                                            command=lambda: toggle_check_all_workers(evening_checkboxes))
toggle_check_all_evening_button.grid(row=4, column=1, padx=pad_x, pady=pad_y)

# Function to create a button to edit worker settings
def create_edit_button(frame, worker, shift, row_num):
    global edit_buttons_dict

    # Check if the edit button for this worker already exists
    if worker not in edit_buttons_dict:
        edit_button = tk.Button(frame, text="Edit", command=lambda w=worker, s=shift: edit_worker_settings(w, s))
        edit_button.grid(row=row_num, column=2, sticky='w')  # Use grid with sticky='w' to align buttons to the left
        edit_buttons_dict[worker] = edit_button  # Store the button in the dictionary

# Function to edit worker settings
def edit_worker_settings(worker_name, shift):
    # Create a new window for editing worker settings
    edit_window = tk.Toplevel(window)
    edit_window.title(f"Edit Settings for {worker_name} ({shift})")

    # Add widgets for editing settings
    tk.Label(edit_window, text=f"Edit settings for {worker_name} ({shift})").pack()

    # Create checkboxes for stations
    stations_label = tk.Label(edit_window, text="Stations:")
    stations_label.pack()

    stations_checkboxes = []
    for station in stations:
        var = tk.BooleanVar(value=(station in workers[shift][worker_name]))
        checkbox = tk.Checkbutton(edit_window, text=station, variable=var)
        checkbox.pack(anchor=tk.W)
        stations_checkboxes.append((var, station))

    # Create checkboxes for shift time options
    shift_label = tk.Label(edit_window, text="Shift Time:")
    shift_label.pack()

    shift_checkboxes = []
    shift_options = intervals_day_shift if shift == 'Day Shift' else intervals_evening_shift
    for option in shift_options:
        var = tk.BooleanVar(value=(option in shift_options))
        checkbox = tk.Checkbutton(edit_window, text=option, variable=var)
        checkbox.pack(anchor=tk.W)
        shift_checkboxes.append((var, option))

    # Function to save edited settings
    def save_settings():
        # Update worker data with edited stations and shift time
        selected_stations = [station for var, station in stations_checkboxes if var.get()]
        workers[shift][worker_name] = selected_stations

        selected_shift_options = [option for var, option in shift_checkboxes if var.get()]
        if shift == 'Day Shift':
            global intervals_day_shift
            intervals_day_shift = selected_shift_options
        else:
            global intervals_evening_shift
            intervals_evening_shift = selected_shift_options

        # Save updated workers data
        save_workers_data()

        # Close the edit window
        edit_window.destroy()
        # Update checkboxes and assigned data display
        on_checkbox_select()

    # Button to save edited settings
    save_button = tk.Button(edit_window, text="Save", command=save_settings)
    save_button.pack()

    # Button to delete the worker
    delete_button = tk.Button(edit_window, text="Delete", command=lambda: delete_worker(worker_name, shift))
    delete_button.pack()

# Function to delete a worker
def delete_worker(worker_name, shift):
    confirmed = messagebox.askyesno("Delete Worker", f"Are you sure you want to delete {worker_name} ({shift})?")
    if confirmed:
        del workers[shift][worker_name]
        save_workers_data()
        messagebox.showinfo("Success", f"{worker_name} ({shift}) deleted successfully.")

# Update the on_checkbox_select function to update edit buttons after editing worker settings
def on_checkbox_select():
    # Clear any existing buttons
    for widget in window.winfo_children():
        if isinstance(widget, tk.Button) and widget.winfo_parent() in (day_worker_frame, evening_worker_frame):
            widget.destroy()


# Call on_checkbox_select initially to populate edit buttons
on_checkbox_select()

# Populate checkboxes for day shift workers
for i, worker in enumerate(workers['Day Shift']):
    var = tk.BooleanVar()
    checkbox = tk.Checkbutton(day_worker_frame, text=worker, variable=var, onvalue=True, offvalue=False, command=on_checkbox_select)
    checkbox.grid(row=i, column=0, sticky='w')  # Use grid with sticky='w' to align checkboxes to the left
    create_edit_button(day_worker_frame, worker, 'Day Shift', i)  # Pass the row number 'i'
    day_checkboxes.append(var)

# Populate checkboxes for evening shift workers
for i, worker in enumerate(workers['Evening Shift']):
    var = tk.BooleanVar()
    checkbox = tk.Checkbutton(evening_worker_frame, text=worker, variable=var, onvalue=True, offvalue=False, command=on_checkbox_select)
    checkbox.grid(row=i, column=0, sticky='w')  # Use grid with sticky='w' to align checkboxes to the left
    create_edit_button(evening_worker_frame, worker, 'Evening Shift', i)  # Pass the row number 'i'
    evening_checkboxes.append(var)


# Shift selection buttons with colors
shift_var = tk.StringVar()


# Create a StringVar to update the assigned workers dynamically
assigned_workers_var = tk.StringVar()


# Function to create a new worker
def create_worker():
    # Create a new window for creating a worker
    new_window = tk.Toplevel(window)
    new_window.title("Create Worker")

    new_window.geometry("300x400")  # Set your desired width and height here

    # Label and entry for worker's name
    tk.Label(new_window, text="Enter worker's name:").pack()
    worker_name_entry = tk.Entry(new_window)
    worker_name_entry.pack()

    # Radiobuttons to select shift
    tk.Label(new_window, text="Select shift:").pack()
    shift_var = tk.StringVar(value="Day Shift")
    tk.Radiobutton(new_window, text="Day Shift", variable=shift_var, value="Day Shift").pack()
    tk.Radiobutton(new_window, text="Evening Shift", variable=shift_var, value="Evening Shift").pack()

    # Choose stations from a listbox
    tk.Label(new_window, text="Choose stations:").pack()
    stations_var = tk.StringVar(value=stations)
    stations_listbox = tk.Listbox(new_window, listvariable=stations_var, selectmode=tk.EXTENDED, height=10)
    stations_listbox.pack()

    # Function to handle creating the worker after input validation
    def create_worker_action():
        worker_name = worker_name_entry.get()
        shift = shift_var.get()
        selected_stations = [stations_listbox.get(idx) for idx in stations_listbox.curselection()]

        # Check if all required fields are filled
        if not worker_name:
            messagebox.showerror("Error", "Please enter worker's name.")
            return
        if not selected_stations:
            messagebox.showerror("Error", "Please select at least one station.")
            return

        # Update the workers data dictionary with the new worker's information
        if shift not in workers:
            workers[shift] = {}
        workers[shift][worker_name] = selected_stations

        # Update the UI to display the new worker in the list of workers
        if shift == 'Day Shift':
            var = tk.BooleanVar()
            checkbox = tk.Checkbutton(day_worker_frame, text=worker_name, variable=var, onvalue=True, offvalue=False,
                                      command=on_checkbox_select)
            checkbox.grid(row=len(day_checkboxes), column=0, sticky='w')
            create_edit_button(day_worker_frame, worker_name, 'Day Shift', len(day_checkboxes))
            day_checkboxes.append(var)
        else:
            var = tk.BooleanVar()
            checkbox = tk.Checkbutton(evening_worker_frame, text=worker_name, variable=var, onvalue=True,
                                      offvalue=False, command=on_checkbox_select)
            checkbox.grid(row=len(evening_checkboxes), column=0, sticky='w')
            create_edit_button(evening_worker_frame, worker_name, 'Evening Shift', len(evening_checkboxes))
            evening_checkboxes.append(var)

        # Close the create worker window
        new_window.destroy()

        # Show success message
        messagebox.showinfo("Success", "Worker created successfully.")

    # Button to create the worker after validation
    create_button = tk.Button(new_window, text="Create Worker", command=create_worker_action)
    create_button.pack(padx=10, pady=20)  # Adjust the values as per your requirements


# Button to create a new worker
create_worker_button = tk.Button(window, text="Create Worker", command=create_worker, bg='lightgrey')
create_worker_button.grid(row=6, column=0, columnspan=1, padx=pad_x, pady=pad_y)
assigned_stations_intervals = {interval: [] for interval in intervals_day_shift + intervals_evening_shift}

# Define a function to check station availability and avoid double assignment
def check_station_availability(station, interval, assigned_stations_intervals):
    return station not in assigned_stations_intervals[interval]

tm_name_label = tk.Label(window, text="Team Manager:", bg="lightgrey", fg="black")
tm_name_label.grid(row=0, column=0, padx=pad_x, pady=pad_y, sticky='e')  # Adjust row and column as needed

tm_name_entry = tk.Entry(window)
tm_name_entry.grid(row=0, column=1, padx=pad_x, pady=pad_y, sticky='w')  # Adjust row and column as needed

back_up_label = tk.Label(window, text="Back up:", bg="lightgrey", fg="black")
back_up_label.grid(row=1, column=0, padx=pad_x, pady=pad_y, sticky='e')  # Adjust row and column as needed

back_up_name_entry = tk.Entry(window)
back_up_name_entry.grid(row=1, column=1, padx=pad_x, pady=pad_y, sticky='w')  # Adjust row and column as needed

late_tm_label = tk.Label(window, text="Sen Vagt - Aften", bg="lightgrey", fg="black")
late_tm_label.grid(row=2, column=0, padx=pad_x, pady=pad_y, sticky='e')  # Adjust row and column as needed

late_tm_entry = tk.Entry(window)
late_tm_entry.grid(row=2, column=1, padx=pad_x, pady=pad_y, sticky='w')  # Adjust row and column as needed



# Function to assign workers to stations and save data to Excel

def assign_workers():
    global assigned_stations_intervals

    # Reset assigned_stations_intervals dictionary
    assigned_stations_intervals = {interval: [] for interval in intervals_day_shift + intervals_evening_shift}

    selected_day_workers = [worker for worker, var in zip(workers['Day Shift'], day_checkboxes) if var.get()]
    selected_evening_workers = [worker for worker, var in zip(workers['Evening Shift'], evening_checkboxes) if var.get()]

    # List of unassigned workers
    unassigned_day_workers = [worker for worker in workers['Day Shift'] if worker not in selected_day_workers]
    unassigned_evening_workers = [worker for worker in workers['Evening Shift'] if worker not in selected_evening_workers]

    # Shuffle the unassigned workers lists to randomize the assignment process
    random.shuffle(unassigned_day_workers)
    random.shuffle(unassigned_evening_workers)

    if not selected_day_workers and not selected_evening_workers:
        messagebox.showerror("Error", "Please select workers.")
        return

    assigned_workers_list = []  # List to store assigned workers
    assigned_stations_per_worker = {}  # Dictionary to track assigned stations for each worker and interval

    # Function to check if a station is already assigned to a worker in a specific interval
    def is_station_assigned(worker, station, interval):
        return assigned_stations_per_worker.get((worker, interval)) == station

    # Process selected workers from the day shift checkboxes
    for worker in selected_day_workers:
        intervals = intervals_day_shift
        assigned_stations_per_worker[worker] = {}  # Initialize assigned stations for the worker
        for interval in intervals:
            available_stations = [s for s in workers['Day Shift'][worker] if
                                  check_station_availability(s, interval, assigned_stations_intervals) and not is_station_assigned(worker, s, interval)]
            if available_stations:
                station = random.choice(available_stations)
                assigned_stations_intervals[interval].append(station)
                assigned_stations_per_worker[worker][interval] = station
                assigned_workers_list.append((worker, station, interval))
            elif unassigned_day_workers:
                # Assign to an unassigned worker if no available stations
                unassigned_worker = unassigned_day_workers.pop()
                assigned_workers_list.append((unassigned_worker, "No station available", interval))
            elif assigned_workers_list:
                # Find an empty station to assign the last worker
                empty_stations = [s for s in stations if check_station_availability(s, interval, assigned_stations_intervals)]
                if empty_stations:
                    station = random.choice(empty_stations)
                    assigned_stations_intervals[interval].append(station)
                    assigned_workers_list.append((worker, station, interval))

    # Process selected workers from the evening shift checkboxes
    for worker in selected_evening_workers:
        intervals = intervals_evening_shift
        assigned_stations_per_worker[worker] = {}  # Initialize assigned stations for the worker
        for interval in intervals:
            available_stations = [s for s in workers['Evening Shift'][worker] if
                                  check_station_availability(s, interval, assigned_stations_intervals) and not is_station_assigned(worker, s, interval)]
            if available_stations:
                station = random.choice(available_stations)
                assigned_stations_intervals[interval].append(station)
                assigned_stations_per_worker[worker][interval] = station
                assigned_workers_list.append((worker, station, interval))
            elif unassigned_evening_workers:
                # Assign to an unassigned worker if no available stations
                unassigned_worker = unassigned_evening_workers.pop()
                assigned_workers_list.append((unassigned_worker, "No station available", interval))
            elif assigned_workers_list:
                # Find an empty station to assign the last worker
                empty_stations = [s for s in stations if check_station_availability(s, interval, assigned_stations_intervals)]
                if empty_stations:
                    station = random.choice(empty_stations)
                    assigned_stations_intervals[interval].append(station)
                    assigned_workers_list.append((worker, station, interval))

    # Get the team manager's, Back up and evening shift name from the entry fields
    team_manager_name = tm_name_entry.get()
    back_up_name = back_up_name_entry.get()
    late_tm_name = late_tm_entry.get()


    # Save the assigned data with the team manager's name
    # Save the assigned data with the team manager's name
    save_assigned_data(assigned_workers_list, team_manager_name, back_up_name, late_tm_name)

    # Update the assigned workers label to display the assigned workers
    assigned_workers_text = "\n".join(
        [f"{worker} - {station} ({interval})" for worker, station, interval in assigned_workers_list])
    assigned_workers_var.set(assigned_workers_text)

    # Check if all workers have been assigned
    if len(selected_day_workers) == len(workers['Day Shift']) and len(selected_evening_workers) == len(workers['Evening Shift']):
        messagebox.showinfo("Success", "All workers have been assigned.")
    else:
        messagebox.showwarning("Warning", "Not all workers have been assigned to stations.")

# Call the updated assign_workers function



# Button to trigger assignment with a colored background
assign_button = tk.Button(window, text="Assign Workers", command=assign_workers, bg='lightgrey')
assign_button.grid(row=8, column=0, columnspan=4, padx=pad_x, pady=pad_y)

# Label to display assigned workers
assigned_label = tk.Label(window, textvariable=assigned_workers_var)
assigned_label.grid(row=9, column=0, columnspan=4, padx=pad_x, pady=pad_y)

# Function to save workers data to file
def save_workers_data():
    with open(workers_file, 'w') as file:
        json.dump(workers, file)
    messagebox.showinfo("Save", "Workers data saved successfully.")


# Save button to save workers data to file
save_button = tk.Button(window, text="Save Workers Data", command=save_workers_data, bg='lightgrey')
save_button.grid(row=8, column=0, columnspan=1, padx=pad_x, pady=pad_y)


# Function to load workers data from file
def load_workers_data():
    try:
        with open(workers_file, 'r') as file:
            workers_data = json.load(file)
            # Clear the worker listbox before adding workers from loaded data




        messagebox.showinfo("Load", "Workers data loaded successfully.")
    except FileNotFoundError:
        messagebox.showwarning("Load", "Workers data file not found.")


# Load button to load workers data from file
load_button = tk.Button(window, text="Load Workers Data", command=load_workers_data, bg='lightgrey')
load_button.grid(row=7, column=0, columnspan=1, padx=pad_x, pady=pad_y)



# Function to save assigned data to Excel file with stations, time intervals, and workers' names
def save_assigned_data(data_list, team_manager_name, back_up_name, late_tm_name  ):
    wb = Workbook()
    ws = wb.active

    # Add the team manager's name in the first row
    ws.append(['Team Manager', team_manager_name])
    ws.append(['Back up', back_up_name])
    ws.append(['Sen vagt - Aften', late_tm_name])



    combined_shifts = intervals_day_shift + intervals_evening_shift

    # Add a row for stations above the station names
    stations_row = ['Week and day of the week', '4:00-6:00'] + combined_shifts
    ws.append(stations_row)

    # Create a dictionary to store assigned workers for each interval and station
    # Ensure combined_shifts is used to maintain the order
    assigned_workers_dict = {interval: {station: [] for station in stations} for interval in combined_shifts}

    # Populate the assigned workers dictionary
    for worker, station, interval in data_list:
        if interval in assigned_workers_dict and station in assigned_workers_dict[interval]:
            assigned_workers_dict[interval][station].append(worker)

    # Iterate through each station and create rows for each station
    for station in stations:
        row_data = [''] * (len(combined_shifts) + 2)  # Adjust for empty and station columns
        row_data[0] = station  # Set the station name in the second column

        # Iterate through each interval in the order of combined_shifts
        for idx, interval in enumerate(combined_shifts):
            if interval in assigned_workers_dict and station in assigned_workers_dict[interval]:
                workers_assigned = assigned_workers_dict[interval][station]
                row_data[idx + 2] = ', '.join(workers_assigned)  # Index +2 because of empty and station columns

        ws.append(row_data)  # Add the completed row to the worksheet

    # Auto-fit columns after adding data
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    assigned_file = 'assigned_data.xlsx'
    wb.save(assigned_file)
    messagebox.showinfo("Save", "Assigned data saved to Excel successfully.")



# Run the main loop
window.mainloop()
