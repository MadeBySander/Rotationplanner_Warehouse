import tkinter as tk
from tkinter import messagebox
import random
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# File paths
workers_file = 'workers_data.json'
assigned_file = 'assigned_data.xlsx'

# Check if workers data file exists
try:
    with open(workers_file, 'r') as file:
        workers = json.load(file)
except FileNotFoundError:
    # Default data if file doesn't exist
    workers = {
        'Day Shift': {
            'Alice': ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job',
                      'Make 301', 'Make 302',
                      'Make 303'],
            'Bob': ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job',
                    'Make 301', 'Make 302',
                    'Make 303'],
            'Emily': ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job',
                      'Make 301', 'Make 302',
                      'Make 303']
        },
        'Evening Shift': {
            'Charlie': ['Make 301', 'Make 302', 'Make 303'],
            'David': ['Make 301', 'Make 302', 'Make 303']
        }
    }

# Sample data
# Sample data
stations = ['Pak 5 (Truck)', 'Pak 5 (Gulv)', 'Pak 1', 'Pak 2', 'Pak 3', 'Pak 4', 'Mezzanin', 'Truck Job', 'Make 301',
            'Make 302', 'Make 303']
intervals_day_shift = ['6-10', '10-14']  # Updated intervals for day shift
intervals_evening_shift = ['14-18', '18-22:30']  # Updated intervals for evening shift

# Create the main window
window = tk.Tk()
window.title("ALFA LAVAL DC - Create Rotationsplan")

# Add padding around the elements
pad_x = 20
pad_y = 10

# Labels for the checkboxes
day_frame_label = tk.Label(window, text="Day Shift Workers")
day_frame_label.grid(row=1, column=0, padx=pad_x, pady=pad_y)

evening_frame_label = tk.Label(window, text="Evening Shift Workers")
evening_frame_label.grid(row=1, column=1, padx=pad_x, pady=pad_y)

# Display the checkboxes with padding
day_worker_frame = tk.Frame(window)
day_worker_frame.grid(row=2, column=0, padx=pad_x, pady=pad_y)

evening_worker_frame = tk.Frame(window)
evening_worker_frame.grid(row=2, column=1, padx=pad_x, pady=pad_y)

day_checkboxes = []
evening_checkboxes = []


# Function to handle checkbox selection
def on_checkbox_select():
    selected_day_workers = [worker.get() for worker in day_checkboxes if worker.get()]
    selected_evening_workers = [worker.get() for worker in evening_checkboxes if worker.get()]
    print("Selected Day Shift Workers:", selected_day_workers)
    print("Selected Evening Shift Workers:", selected_evening_workers)


# Populate checkboxes for day shift workers
for worker in workers['Day Shift']:
    var = tk.BooleanVar()
    checkbox = tk.Checkbutton(day_worker_frame, text=worker, variable=var, onvalue=True, offvalue=False,
                              command=on_checkbox_select)
    checkbox.pack(anchor=tk.W)
    day_checkboxes.append(var)

# Populate checkboxes for evening shift workers
for worker in workers['Evening Shift']:
    var = tk.BooleanVar()
    checkbox = tk.Checkbutton(evening_worker_frame, text=worker, variable=var, onvalue=True, offvalue=False,
                              command=on_checkbox_select)
    checkbox.pack(anchor=tk.W)
    evening_checkboxes.append(var)

# Shift selection buttons with colors
shift_var = tk.StringVar()

# Static stations display with colors
for i, station in enumerate(stations):
    if i < len(stations) // 2:
        bg_color = 'lightyellow' if i % 2 == 0 else 'lightcyan'  # Alternating background colors
        tk.Label(window, text=station, bg=bg_color).grid(row=i + 1, column=3, padx=pad_x, pady=pad_y)
    else:
        j = i - len(stations) // 2
        bg_color = 'lightyellow' if j % 2 == 0 else 'lightcyan'  # Alternating background colors
        tk.Label(window, text=station, bg=bg_color).grid(row=j + 1, column=4, padx=pad_x, pady=pad_y)

# Create a StringVar to update the assigned workers dynamically
assigned_workers_var = tk.StringVar()


# Function to create a new worker
def create_worker():
    new_window = tk.Toplevel(window)
    new_window.title("Create Worker")

    tk.Label(new_window, text="Enter worker's name:").pack()
    worker_name_entry = tk.Entry(new_window)
    worker_name_entry.pack()

    tk.Label(new_window, text="Select shift:").pack()
    shift_var = tk.StringVar(value="Day Shift")
    tk.Radiobutton(new_window, text="Day Shift", variable=shift_var, value="Day Shift").pack()
    tk.Radiobutton(new_window, text="Evening Shift", variable=shift_var, value="Evening Shift").pack()

    tk.Label(new_window, text="Choose stations:").pack()
    stations_var = tk.StringVar(value=stations)
    stations_listbox = tk.Listbox(new_window, listvariable=stations_var, selectmode=tk.MULTIPLE, height=4)
    stations_listbox.pack()


# Add a label and entry field for Team Manager's name
tm_name_label = tk.Label(window, text="Team Manager:")
tm_name_label.grid(row=0, column=0, padx=pad_x, pady=pad_y)
tm_name_entry = tk.Entry(window)
tm_name_entry.grid(row=0, column=1, padx=pad_x, pady=pad_y)

# Button to create a new worker
create_worker_button = tk.Button(window, text="Create Worker", command=create_worker, bg='lightgrey')
create_worker_button.grid(row=3, column=0, columnspan=1, padx=pad_x, pady=pad_y)

assigned_stations_intervals = {interval: [] for interval in intervals_day_shift + intervals_evening_shift}


# Function to assign workers to stations randomly
# Function to assign workers to stations randomly
def assign_workers():
    selected_day_workers = [worker for worker, var in zip(workers['Day Shift'], day_checkboxes) if var.get()]
    selected_evening_workers = [worker for worker, var in zip(workers['Evening Shift'], evening_checkboxes) if
                                var.get()]

    if not selected_day_workers and not selected_evening_workers:
        messagebox.showerror("Error", "Please select workers.")
        return

    assigned_workers_list = []  # List to store assigned workers

    # Process selected workers from the day shift checkboxes
    for worker in selected_day_workers:
        intervals = intervals_day_shift
        for interval in intervals:
            available_stations = [s for s in workers['Day Shift'][worker] if
                                  s not in assigned_stations_intervals[interval]]
            if available_stations:
                station = random.choice(available_stations)
                assigned_stations_intervals[interval].append(station)
                assigned_workers_list.append((worker, station, interval))

    # Process selected workers from the evening shift checkboxes
    for worker in selected_evening_workers:
        intervals = intervals_evening_shift
        for interval in intervals:
            available_stations = [s for s in workers['Evening Shift'][worker] if
                                  s not in assigned_stations_intervals[interval]]
            if available_stations:
                station = random.choice(available_stations)
                assigned_stations_intervals[interval].append(station)
                assigned_workers_list.append((worker, station, interval))

    # Further processing or saving the assigned workers as needed

    team_manager_name = tm_name_entry.get()  # Get the team manager's name
    save_assigned_data(assigned_workers_list, team_manager_name)  # Save the assigned data

    # Update the assigned workers label to display the assigned workers
    assigned_workers_text = "\n".join(
        [f"{worker} - {station} ({interval})" for worker, station, interval in assigned_workers_list])
    assigned_workers_var.set(assigned_workers_text)

    messagebox.showinfo("Success", "Workers assigned successfully.")


# Button to trigger assignment with a colored background
assign_button = tk.Button(window, text="Assign Workers", command=assign_workers, bg='lightgrey')
assign_button.grid(row=8, column=0, columnspan=4, padx=pad_x, pady=pad_y)

# Label to display assigned workers
assigned_label = tk.Label(window, textvariable=assigned_workers_var)
assigned_label.grid(row=9, column=0, columnspan=4, padx=pad_x, pady=pad_y)


# Function to save workers data to file
def save_workers_data():
    with open(workers_file, 'w') as file:
        json.dump(workers, file)
    messagebox.showinfo("Save", "Workers data saved successfully.")


# Save button to save workers data to file
save_button = tk.Button(window, text="Save Workers Data", command=save_workers_data, bg='lightgrey')
save_button.grid(row=4, column=0, columnspan=1, padx=pad_x, pady=pad_y)


# Function to load workers data from file
def load_workers_data():
    try:
        with open(workers_file, 'r') as file:
            workers_data = json.load(file)
            # Clear the worker listbox before adding workers from loaded data

        messagebox.showinfo("Load", "Workers data loaded successfully.")
    except FileNotFoundError:
        messagebox.showwarning("Load", "Workers data file not found.")


# Load button to load workers data from file
load_button = tk.Button(window, text="Load Workers Data", command=load_workers_data, bg='lightgrey')
load_button.grid(row=5, column=0, columnspan=1, padx=pad_x, pady=pad_y)


# Function to save assigned data to Excel file with stations, time intervals, and workers' names
# Function to save assigned data to Excel file with stations, time intervals, and workers' names
def save_assigned_data(data_list, team_manager_name):
    wb = Workbook()
    ws = wb.active

    # Add the team manager's name in the first row
    ws.append(['Team Manager', team_manager_name])

    # Add a row for stations above the station names
    stations_row = ['', 'Stations'] + ['6-10', '10-14', '14-18', '18-22:30']
    ws.append(stations_row)

    # Create a dictionary to store assigned workers for each interval and station
    assigned_workers_dict = {interval: {station: [] for station in stations} for interval in
                             intervals_day_shift + intervals_evening_shift}

    # Populate the assigned workers dictionary
    for worker, station, interval in data_list:
        if interval in assigned_workers_dict and station in assigned_workers_dict[interval]:
            assigned_workers_dict[interval][station].append(worker)
        else:
            messagebox.showerror("Error", f"Invalid station or interval: {station}, {interval}")

    # Iterate through each station and create rows for each station
    for station in stations:
        row_data = [''] * (len(intervals_day_shift + intervals_evening_shift) + 2)  # Initialize row with empty values
        row_data[1] = station  # Set the station name in the second column

        # Iterate through each time interval
        for interval in ['6-10', '10-14', '14-18', '18-22:30']:
            if interval in assigned_workers_dict and station in assigned_workers_dict[interval]:
                workers_assigned = assigned_workers_dict[interval][
                    station]  # Get assigned workers for this interval and station
                if interval in intervals_day_shift:
                    row_data[intervals_day_shift.index(interval) + 2] = ', '.join(workers_assigned)
                elif interval in intervals_evening_shift:
                    row_data[intervals_evening_shift.index(interval) + len(intervals_day_shift) + 2] = ', '.join(
                        workers_assigned)
            else:
                if interval in intervals_day_shift:
                    row_data[intervals_day_shift.index(interval) + 2] = ''
                elif interval in intervals_evening_shift:
                    row_data[intervals_evening_shift.index(interval) + len(intervals_day_shift) + 2] = ''

        ws.append(row_data)  # Add the completed row_data to the worksheet

    # Auto-fit columns after adding data
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    assigned_file = 'assigned_data.xlsx'
    wb.save(assigned_file)
    messagebox.showinfo("Save", "Assigned data saved to Excel successfully.")


# Save button to save assigned data to Excel file
save_to_excel_button = tk.Button(window, text="Save to Excel",
                                 command=lambda: save_assigned_data(assigned_workers_var.get().split('\n'),
                                                                    tm_name_entry.get()),
                                 bg='lightgrey')
save_to_excel_button.grid(row=4, column=1, columnspan=1, padx=pad_x, pady=pad_y)

# Run the main loop
window.mainloop()